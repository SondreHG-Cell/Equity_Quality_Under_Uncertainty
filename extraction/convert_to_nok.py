"""
FX conversion: convert local-currency LSEG outputs to NOK using Norges Bank
exchange rates.

Inputs:
    instrument_currencies.csv
    total_returns_local.csv
    historical_market_cap_local.csv   optional

FX inputs:
    NOK_EUR.xlsx, NOK_USD.xlsx, NOK_SEK.xlsx, NOK_DKK.xlsx, NOK_ISK.xlsx

Outputs:
    total_returns_nok.csv
    total_returns_nok_monthly_returns.csv
    historical_market_cap_nok.csv
    fx_conversion_audit.csv
"""

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# =============================================================================
# Configuration
# =============================================================================

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data"
FX_DATA = HERE

FX_FILES = {
    "EUR": FX_DATA / "NOK_EUR.xlsx",
    "USD": FX_DATA / "NOK_USD.xlsx",
    "SEK": FX_DATA / "NOK_SEK.xlsx",
    "DKK": FX_DATA / "NOK_DKK.xlsx",
    "ISK": FX_DATA / "NOK_ISK.xlsx",
}


# =============================================================================
# Norges Bank FX reader
# =============================================================================

def read_norges_bank_fx(path: Path) -> tuple[pd.DataFrame, str, str]:
    """
    Read a Norges Bank EXR Excel export.

    Returns:
        (df, base_cur, quote_cur)
        df has columns ['date', 'rate', 'rate_per_unit']
        rate_per_unit = rate / 10**UNIT_MULT
            so 1 unit of base_cur = rate_per_unit units of quote_cur.

    For our use case quote_cur is always NOK, so:
        NOK_amount = local_amount * rate_per_unit
    """
    if not path.exists():
        raise FileNotFoundError(f"FX file not found: {path}")

    wb = load_workbook(path, data_only=True)

    if "Dataset" not in wb.sheetnames:
        raise ValueError(f"{path} does not have a 'Dataset' sheet.")

    ws = wb["Dataset"]

    # Read metadata from rows 1-20.
    meta = {}

    for row in range(1, 21):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value

        if key is not None and val is not None:
            meta[str(key).strip()] = val

    base_cur = str(meta.get("BASE_CUR", "")).strip().upper()
    quote_cur = str(meta.get("QUOTE_CUR", "")).strip().upper()

    try:
        unit_mult = int(meta.get("UNIT_MULT", 0))
    except (TypeError, ValueError):
        unit_mult = 0

    if quote_cur != "NOK":
        raise ValueError(
            f"{path} has QUOTE_CUR={quote_cur}, expected NOK. "
            "The FX files must be quoted as NOK per unit of foreign currency."
        )

    # Read dates from row 22 and rates from row 23.
    dates = []
    rates = []

    for col in range(1, ws.max_column + 1):
        date_value = ws.cell(row=22, column=col).value
        rate_value = ws.cell(row=23, column=col).value

        if date_value is not None and rate_value is not None:
            dates.append(date_value)
            rates.append(rate_value)

    df = pd.DataFrame({
        "date": pd.to_datetime(dates, errors="coerce"),
        "rate": pd.to_numeric(rates, errors="coerce"),
    })

    df = df.dropna(subset=["date", "rate"])
    df = df.sort_values("date").reset_index(drop=True)

    # Respect UNIT_MULT, e.g. ISK may be quoted per 100 ISK.
    df["rate_per_unit"] = df["rate"] / (10 ** unit_mult)

    return df, base_cur, quote_cur

def load_all_fx() -> dict[str, pd.DataFrame | None]:
    """
    Load all FX series.

    Returns:
        dict currency -> DataFrame with columns date, rate

    NOK is stored as None because NOK to NOK is identity.
    """
    fx = {}

    print(f"\nLoading FX rates from {FX_DATA}:")

    for ccy, path in FX_FILES.items():
        if not path.exists():
            print(f"  {ccy}: SKIP, file not found at {path}")
            continue

        df, base_cur, quote_cur = read_norges_bank_fx(path)

        if base_cur != ccy:
            print(
                f"  WARNING: {path.name} BASE_CUR={base_cur}, expected {ccy}. "
                "Using filename currency."
            )

        fx[ccy] = (
            df[["date", "rate_per_unit"]]
            .rename(columns={"rate_per_unit": "rate"})
            .copy()
        )

        print(
            f"  {ccy}: {len(fx[ccy]):,} observations from "
            f"{fx[ccy]['date'].min().date()} to {fx[ccy]['date'].max().date()}"
        )

    # Identity conversion.
    fx["NOK"] = None

    return fx


# =============================================================================
# FX lookup helpers
# =============================================================================

def get_month_end_fx(fx_daily: pd.DataFrame) -> pd.Series:
    """
    Compute month-end FX (last available daily rate within each calendar month).

    Returns a Series indexed by YYYY-MM strings.
    """
    df = fx_daily.set_index("date").sort_index()
    monthly = df["rate"].resample("ME").last()
    monthly.index = monthly.index.strftime("%Y-%m")
    return monthly


def build_month_end_fx_levels(
    fx: dict[str, pd.DataFrame | None],
    months: list[str],
) -> dict[str, pd.Series | None]:
    """
    Build month-end FX level series for each currency.

    Keeps one extra previous month so the first stock-return month can use:
        FX_t / FX_{t-1} - 1

    FX level means:
        NOK per 1 unit of local currency
    """
    out = {}

    month_periods = pd.PeriodIndex(months, freq="M")
    first_lag_month = (month_periods.min() - 1).strftime("%Y-%m")
    months_with_lag = [first_lag_month] + list(months)

    for ccy, df_daily in fx.items():
        if ccy == "NOK":
            out[ccy] = None
            continue

        if df_daily is None:
            continue

        monthly_fx = get_month_end_fx(df_daily)
        out[ccy] = monthly_fx.reindex(months_with_lag)

    return out


def build_monthly_fx_returns(
    fx_levels: dict[str, pd.Series | None],
    months: list[str],
) -> dict[str, pd.Series]:
    """
    Build monthly FX return series from month-end FX levels.

    Formula:
        FX_RETURN_t = FX_LEVEL_t / FX_LEVEL_{t-1} - 1

    For NOK:
        FX_RETURN = 0
    """
    out = {}

    for ccy, level_series in fx_levels.items():
        if ccy == "NOK":
            out[ccy] = pd.Series(0.0, index=months)
            continue

        if level_series is None:
            out[ccy] = pd.Series(np.nan, index=months)
            continue

        fx_return = level_series.pct_change()

        # Keep only the months matching the stock-return file.
        out[ccy] = fx_return.reindex(months)

    return out


# =============================================================================
# Total return conversion
# =============================================================================

def convert_total_returns_to_nok(
    total_returns_local_csv: Path,
    currencies: pd.DataFrame,
    fx: dict[str, pd.DataFrame | None],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convert monthly local total returns to monthly NOK total returns.

    Input:
        rows = tickers
        columns = YYYY-MM
        values = monthly local total returns

    Formula:
        TR_NOK = (1 + TR_LOCAL) * (1 + FX_RETURN) - 1

    where:
        FX_RETURN = monthly percentage change in NOK per local currency.
    """
    df_local = pd.read_csv(total_returns_local_csv, index_col="Ticker")
    df_local = df_local.apply(pd.to_numeric, errors="coerce")

    print(
        f"\nConverting total returns: "
        f"{df_local.shape[0]} tickers × {df_local.shape[1]} months"
    )

    currencies = currencies.copy()
    currencies["Ticker"] = currencies["Ticker"].astype(str).str.strip()
    currencies["Currency"] = currencies["Currency"].astype(str).str.strip().str.upper()

    ccy_map = currencies.set_index("Ticker")["Currency"].to_dict()

    months = list(df_local.columns)

    fx_levels = build_month_end_fx_levels(fx, months)
    fx_returns = build_monthly_fx_returns(fx_levels, months)

    converted = df_local.copy()

    audit_rows = []

    n_no_currency = 0
    n_unknown_currency = 0
    n_unsupported_currency = 0
    n_missing_fx_cells = 0

    for ticker in df_local.index:
        ccy = ccy_map.get(ticker)

        if ccy is None or ccy == "" or pd.isna(ccy):
            converted.loc[ticker, :] = np.nan
            n_no_currency += 1
            continue

        ccy = str(ccy).strip().upper()

        if ccy in {"UNKNOWN", "NAN"}:
            converted.loc[ticker, :] = np.nan
            n_unknown_currency += 1
            continue

        if ccy == "NOK":
            # Local return is already NOK return.
            continue

        if ccy not in fx_returns:
            converted.loc[ticker, :] = np.nan
            n_unsupported_currency += 1
            continue

        fx_ret_series = fx_returns[ccy]

        for month in months:
            local_ret = df_local.loc[ticker, month]

            if pd.isna(local_ret):
                continue

            fx_ret = fx_ret_series.loc[month] if month in fx_ret_series.index else np.nan

            if pd.isna(fx_ret):
                converted.loc[ticker, month] = np.nan
                n_missing_fx_cells += 1
            else:
                converted.loc[ticker, month] = (1.0 + local_ret) * (1.0 + fx_ret) - 1.0

    # Build audit sample.
    sample_months = []

    if len(months) > 0:
        sample_months.append(months[0])

    if len(months) > 1:
        sample_months.append(months[len(months) // 2])

    if len(months) > 2:
        sample_months.append(months[-1])

    seen_currencies = set()

    for ticker in df_local.index:
        ccy = ccy_map.get(ticker)

        if ccy is None or pd.isna(ccy):
            continue

        ccy = str(ccy).strip().upper()

        if ccy in seen_currencies:
            continue

        for month in sample_months:
            local_ret = df_local.loc[ticker, month]

            if pd.isna(local_ret):
                continue

            if ccy == "NOK":
                fx_level = 1.0
                fx_ret = 0.0
                nok_ret = local_ret
            else:
                fx_level_series = fx_levels.get(ccy)
                fx_ret_series = fx_returns.get(ccy)

                fx_level = (
                    fx_level_series.loc[month]
                    if fx_level_series is not None and month in fx_level_series.index
                    else np.nan
                )

                fx_ret = (
                    fx_ret_series.loc[month]
                    if fx_ret_series is not None and month in fx_ret_series.index
                    else np.nan
                )

                nok_ret = (
                    (1.0 + local_ret) * (1.0 + fx_ret) - 1.0
                    if pd.notna(fx_ret)
                    else np.nan
                )

            audit_rows.append({
                "Source": "total_returns",
                "Ticker": ticker,
                "Month": month,
                "Currency": ccy,
                "LocalReturn": local_ret,
                "FXLevel_EoM": fx_level,
                "FXReturn": fx_ret,
                "NOKReturn": nok_ret,
            })

        seen_currencies.add(ccy)

    audit = pd.DataFrame(audit_rows)

    print(f"  Tickers with no currency:        {n_no_currency}")
    print(f"  Tickers with UNKNOWN currency:   {n_unknown_currency}")
    print(f"  Tickers with unsupported ccy:    {n_unsupported_currency}")
    print(f"  Cells with missing FX returns:   {n_missing_fx_cells}")

    return converted, audit


# =============================================================================
# Level conversion: market cap
# =============================================================================

def convert_wide_monthly_levels_to_nok(
    local_csv: Path,
    currencies: pd.DataFrame,
    fx: dict[str, pd.DataFrame | None],
    label: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convert wide monthly local-currency levels to NOK.

    This is appropriate for:
        market cap
        price levels
        accounting levels

    Formula:
        VALUE_NOK = VALUE_LOCAL * FX_LEVEL
    """
    df_local = pd.read_csv(local_csv, index_col="Ticker")
    df_local = df_local.apply(pd.to_numeric, errors="coerce")

    print(
        f"\nConverting {label}: "
        f"{df_local.shape[0]} tickers × {df_local.shape[1]} months"
    )

    currencies = currencies.copy()
    currencies["Ticker"] = currencies["Ticker"].astype(str).str.strip()
    currencies["Currency"] = currencies["Currency"].astype(str).str.strip().str.upper()

    ccy_map = currencies.set_index("Ticker")["Currency"].to_dict()

    months = list(df_local.columns)
    fx_levels = build_month_end_fx_levels(fx, months)

    converted = df_local.copy()

    audit_rows = []

    n_no_currency = 0
    n_unknown_currency = 0
    n_unsupported_currency = 0
    n_missing_fx_cells = 0

    for ticker in df_local.index:
        ccy = ccy_map.get(ticker)

        if ccy is None or ccy == "" or pd.isna(ccy):
            converted.loc[ticker, :] = np.nan
            n_no_currency += 1
            continue

        ccy = str(ccy).strip().upper()

        if ccy in {"UNKNOWN", "NAN"}:
            converted.loc[ticker, :] = np.nan
            n_unknown_currency += 1
            continue

        if ccy == "NOK":
            # Already NOK.
            continue

        if ccy not in fx_levels:
            converted.loc[ticker, :] = np.nan
            n_unsupported_currency += 1
            continue

        fx_level_series = fx_levels[ccy]

        if fx_level_series is None:
            converted.loc[ticker, :] = np.nan
            n_unsupported_currency += 1
            continue

        for month in months:
            local_value = df_local.loc[ticker, month]

            if pd.isna(local_value):
                continue

            fx_level = (
                fx_level_series.loc[month]
                if month in fx_level_series.index
                else np.nan
            )

            if pd.isna(fx_level):
                converted.loc[ticker, month] = np.nan
                n_missing_fx_cells += 1
            else:
                converted.loc[ticker, month] = local_value * fx_level

    # Build audit sample.
    sample_months = []

    if len(months) > 0:
        sample_months.append(months[0])

    if len(months) > 1:
        sample_months.append(months[len(months) // 2])

    if len(months) > 2:
        sample_months.append(months[-1])

    seen_currencies = set()

    for ticker in df_local.index:
        ccy = ccy_map.get(ticker)

        if ccy is None or pd.isna(ccy):
            continue

        ccy = str(ccy).strip().upper()

        if ccy in seen_currencies:
            continue

        for month in sample_months:
            local_value = df_local.loc[ticker, month]

            if pd.isna(local_value):
                continue

            if ccy == "NOK":
                fx_level = 1.0
                nok_value = local_value
            else:
                fx_level_series = fx_levels.get(ccy)

                fx_level = (
                    fx_level_series.loc[month]
                    if fx_level_series is not None and month in fx_level_series.index
                    else np.nan
                )

                nok_value = (
                    local_value * fx_level
                    if pd.notna(fx_level)
                    else np.nan
                )

            audit_rows.append({
                "Source": label,
                "Ticker": ticker,
                "Month": month,
                "Currency": ccy,
                "LocalValue": local_value,
                "FXLevel_EoM": fx_level,
                "NOKValue": nok_value,
            })

        seen_currencies.add(ccy)

    audit = pd.DataFrame(audit_rows)

    print(f"  Tickers with no currency:        {n_no_currency}")
    print(f"  Tickers with UNKNOWN currency:   {n_unknown_currency}")
    print(f"  Tickers with unsupported ccy:    {n_unsupported_currency}")
    print(f"  Cells with missing FX levels:    {n_missing_fx_cells}")

    return converted, audit


# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 70)
    print("FX CONVERSION: local currency -> NOK")
    print("=" * 70)

    # -------------------------------------------------------------------------
    # 1. Load currency metadata
    # -------------------------------------------------------------------------

    ccy_path = DATA / "instrument_currencies.csv"

    if not ccy_path.exists():
        raise FileNotFoundError(f"Missing {ccy_path}")

    currencies = pd.read_csv(ccy_path)
    currencies["Ticker"] = currencies["Ticker"].astype(str).str.strip()
    currencies["Currency"] = currencies["Currency"].astype(str).str.strip().str.upper()

    print(f"\nLoaded {len(currencies)} ticker currencies.")
    print(currencies["Currency"].value_counts(dropna=False))

    # -------------------------------------------------------------------------
    # 2. Load FX series
    # -------------------------------------------------------------------------

    fx = load_all_fx()

    needed_currencies = set(currencies["Currency"]) - {"NOK", "UNKNOWN", "NAN", ""}
    available_currencies = {
        ccy for ccy, df in fx.items()
        if ccy == "NOK" or df is not None
    }

    missing_fx = needed_currencies - available_currencies

    if missing_fx:
        print(f"\nWARNING: missing FX series for currencies: {missing_fx}")
        print("Tickers in those currencies will become NaN in NOK output.")

    # -------------------------------------------------------------------------
    # 3. Convert monthly total returns
    # -------------------------------------------------------------------------

    tr_local_path = DATA / "total_returns_local.csv"

    if tr_local_path.exists():
        total_returns_nok, audit_tr = convert_total_returns_to_nok(
            total_returns_local_csv=tr_local_path,
            currencies=currencies,
            fx=fx,
        )

        tr_nok_path = DATA / "total_returns_nok.csv"
        total_returns_nok.to_csv(tr_nok_path)

        print(f"\nSaved {tr_nok_path.name}: {total_returns_nok.shape}")

    else:
        print(f"\nSKIP total returns: {tr_local_path} not found")
        audit_tr = pd.DataFrame()

    # -------------------------------------------------------------------------
    # 4. Convert market cap, if available
    # -------------------------------------------------------------------------

    mktcap_local_path = DATA / "historical_market_cap_local.csv"

    if mktcap_local_path.exists():
        market_cap_nok, audit_mktcap = convert_wide_monthly_levels_to_nok(
            local_csv=mktcap_local_path,
            currencies=currencies,
            fx=fx,
            label="market_cap",
        )

        mktcap_nok_path = DATA / "historical_market_cap_nok.csv"
        market_cap_nok.to_csv(mktcap_nok_path)

        print(f"\nSaved {mktcap_nok_path.name}: {market_cap_nok.shape}")

    else:
        print(f"\nSKIP market cap: {mktcap_local_path} not found")
        audit_mktcap = pd.DataFrame()

    # -------------------------------------------------------------------------
    # 5. Save audit file
    # -------------------------------------------------------------------------

    audit_combined = pd.concat(
        [audit_tr, audit_mktcap],
        ignore_index=True,
    )

    audit_path = DATA / "fx_conversion_audit.csv"
    audit_combined.to_csv(audit_path, index=False)

    print(f"\nSaved audit sample to {audit_path}")

    print("\nAudit interpretation:")
    print("  For total returns:")
    print("      NOKReturn = (1 + LocalReturn) × (1 + FXReturn) - 1")
    print("  For market cap:")
    print("      NOKValue = LocalValue × FXLevel_EoM")

    print("\nDone.")

if __name__ == "__main__":
    main()