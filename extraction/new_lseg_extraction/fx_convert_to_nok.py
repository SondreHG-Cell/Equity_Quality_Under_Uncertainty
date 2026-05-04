"""
FX conversion: convert local-currency LSEG outputs to NOK using Norges Bank
exchange rates.

Inputs (in RAW_DATA directory, produced by extract_lseg.py):
    instrument_currencies.csv         Ticker -> Currency
    all_stock_prices_local.csv        Wide: rows=tickers, cols=YYYY-MM
    historical_market_cap_local.csv   Wide: same shape
    dividends_raw_long_local.csv      Long dividend events. DPS_LOCAL is split-
                                      adjusted gross DPS in Currency.

FX inputs (in FX_DATA directory):
    NOK_EUR.xlsx, NOK_USD.xlsx, NOK_SEK.xlsx, NOK_DKK.xlsx, NOK_ISK.xlsx,
    plus any extra dividend currencies listed in FX_FILES below.
    Each: a Norges Bank EXR Excel export with metadata in rows 1-20,
          dates in row 22, daily rates in row 23.

Outputs (in PROCESSED_DATA directory):
    all_stock_prices_nok.csv          Wide: rows=tickers, cols=YYYY-MM, NOK
    historical_market_cap_nok.csv     Wide: same shape, NOK
    dividends_raw_long_nok.csv        Long: preserves dividend audit columns,
                                      adds FXRate, FXDate, FXSource, DPS_NOK
    dividends_monthly_nok.csv         Wide: rows=tickers, cols=YYYY-MM, NOK
    fx_conversion_audit.csv           Spot-check sample for verification

Conventions
-----------
- Prices and market cap use MONTH-END FX (last business day rate in the month).
- Dividends use EX-DATE FX (last business day rate <= ex-date).
- All FX rates are read from Norges Bank with the UNIT_MULT field respected
  (e.g. ISK is quoted per 100 ISK).
- Rate semantics: NOK = LOCAL × rate (the file gives "NOK per 1 unit of base").
"""

from pathlib import Path
import shutil

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# =============================================================================
# Configuration
# =============================================================================

HERE = Path(__file__).resolve().parent
BASE = HERE.parents[1]
RAW_DATA = BASE / "data" / "raw_data_lseg"
PROCESSED_DATA = BASE / "data" / "processed_data_lseg"
FX_DATA = HERE
MONTHLY_FX_FALLBACK = RAW_DATA / "fx_rates.csv"
OUTPUT_START_DATE = pd.Timestamp("2005-01-01")
OUTPUT_START_MONTH = OUTPUT_START_DATE.strftime("%Y-%m")

# Map from currency code to Norges Bank file name (in the same folder).
FX_FILES = {
    "EUR": FX_DATA / "NOK_EUR.xlsx",
    "USD": FX_DATA / "NOK_USD.xlsx",
    "SEK": FX_DATA / "NOK_SEK.xlsx",
    "DKK": FX_DATA / "NOK_DKK.xlsx",
    "ISK": FX_DATA / "NOK_ISK.xlsx",
    "GBP": FX_DATA / "NOK_GBP.xlsx",
    "PLN": FX_DATA / "NOK_PLN.xlsx",
    "CAD": FX_DATA / "NOK_CAD.xlsx",
    "CHF": FX_DATA / "NOK_CHF.xlsx",
}


# =============================================================================
# Norges Bank FX reader
# =============================================================================

def read_norges_bank_fx(path: Path) -> tuple[pd.DataFrame, str, str]:
    """
    Read a Norges Bank EXR Excel export.

    Returns:
        (df, base_cur, quote_cur)
        df has columns ['date', 'rate', 'rate_per_unit']
        rate_per_unit = rate / 10**UNIT_MULT
            so 1 unit of base_cur = rate_per_unit units of quote_cur.

    For our use case quote_cur is always NOK, so:
        NOK_amount = local_amount * rate_per_unit
    """
    if not path.exists():
        raise FileNotFoundError(f"FX file not found: {path}")

    wb = load_workbook(path, data_only=True)
    if "Dataset" not in wb.sheetnames:
        raise ValueError(f"{path} does not have a 'Dataset' sheet.")
    ws = wb["Dataset"]

    # Read metadata from rows 1-20.
    meta = {}
    for row in range(1, 21):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if key is not None and val is not None:
            meta[str(key).strip()] = val

    base_cur = str(meta.get("BASE_CUR", "")).strip().upper()
    quote_cur = str(meta.get("QUOTE_CUR", "")).strip().upper()
    try:
        unit_mult = int(meta.get("UNIT_MULT", 0))
    except (TypeError, ValueError):
        unit_mult = 0

    if quote_cur != "NOK":
        raise ValueError(
            f"{path} has QUOTE_CUR={quote_cur}, expected NOK. "
            "Norges Bank rates must be quoted as NOK per unit of foreign currency."
        )

    # Read dates from row 22 and rates from row 23.
    last_col = ws.max_column
    dates = []
    rates = []
    for c in range(1, last_col + 1):
        d = ws.cell(row=22, column=c).value
        r = ws.cell(row=23, column=c).value
        if d is not None and r is not None:
            dates.append(d)
            rates.append(r)

    df = pd.DataFrame({
        "date": pd.to_datetime(dates, errors="coerce"),
        "rate": pd.to_numeric(rates, errors="coerce"),
    })
    df = df.dropna().sort_values("date").reset_index(drop=True)
    df["rate_per_unit"] = df["rate"] / (10 ** unit_mult)

    return df, base_cur, quote_cur


def load_all_fx() -> dict[str, pd.DataFrame]:
    """Load all FX series. Returns dict ccy -> df with date and rate_per_unit."""
    fx = {}
    print(f"\nLoading FX rates from {FX_DATA}:")
    for ccy, path in FX_FILES.items():
        if not path.exists():
            print(f"  {ccy}: SKIP (file not found at {path})")
            continue
        df, base, quote = read_norges_bank_fx(path)
        if base != ccy:
            print(f"  WARNING: {path} BASE_CUR={base}, expected {ccy}. Using filename.")
        fx[ccy] = df[["date", "rate_per_unit"]].rename(columns={"rate_per_unit": "rate"})
        print(
            f"  {ccy}: {len(fx[ccy]):,} obs from "
            f"{fx[ccy]['date'].min().date()} to {fx[ccy]['date'].max().date()}"
        )

    # NOK -> NOK is identity, no file needed.
    fx["NOK"] = None
    return fx


def load_monthly_fx_fallback(path: Path) -> dict[str, pd.DataFrame]:
    """
    Load optional monthly FX fallback rates.

    This is only used where the preferred Norges Bank daily series has no
    available observation. The expected shape is the legacy fx_rates.csv:
        Date,SEK,DKK,EUR,ISK,NOK
    with rates already expressed as NOK per one unit of local currency.
    """
    if not path.exists():
        print(f"\nMonthly FX fallback not found at {path}; continuing without fallback.")
        return {}

    df = pd.read_csv(path)
    if "Date" not in df.columns:
        print(f"\nMonthly FX fallback at {path} has no Date column; ignoring.")
        return {}

    df["date"] = pd.to_datetime(df["Date"], errors="coerce") + pd.offsets.MonthEnd(0)
    df = df.dropna(subset=["date"]).sort_values("date")

    fallback = {}
    for ccy in [c for c in df.columns if c not in {"Date", "date"}]:
        ccy_norm = str(ccy).strip().upper()
        rates = pd.to_numeric(df[ccy], errors="coerce")
        sub = pd.DataFrame({"date": df["date"], "rate": rates}).dropna()
        if not sub.empty:
            fallback[ccy_norm] = sub.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)

    if fallback:
        covered = ", ".join(sorted(fallback))
        print(f"\nLoaded monthly FX fallback from {path}: {covered}")
    return fallback


# =============================================================================
# FX lookup helpers
# =============================================================================

def get_month_end_fx(fx_daily: pd.DataFrame) -> pd.Series:
    """
    Compute month-end FX (last available daily rate within each calendar month).

    Returns a Series indexed by YYYY-MM strings.
    """
    df = fx_daily.set_index("date").sort_index()
    monthly = df["rate"].resample("ME").last()
    monthly.index = monthly.index.strftime("%Y-%m")
    return monthly


def get_rate_for_date(
    target_date: pd.Timestamp,
    fx_daily: pd.DataFrame,
) -> tuple[float, pd.Timestamp]:
    """
    Get the FX rate effective on or before target_date (last business day rate
    on or before the date - handles weekends and holidays).

    Returns (rate, date_used).
    """
    sub = fx_daily.loc[fx_daily["date"] <= target_date]
    if sub.empty:
        return np.nan, pd.NaT
    last = sub.iloc[-1]
    return float(last["rate"]), last["date"]


def get_fallback_rate_for_date(
    target_date: pd.Timestamp,
    fx_monthly: pd.DataFrame,
) -> tuple[float, pd.Timestamp]:
    """Get the latest fallback monthly FX rate on or before target_date."""
    sub = fx_monthly.loc[fx_monthly["date"] <= target_date]
    if sub.empty:
        return np.nan, pd.NaT
    last = sub.iloc[-1]
    return float(last["rate"]), last["date"]


# =============================================================================
# Conversion functions
# =============================================================================

def convert_wide_monthly_to_nok(
    local_csv: Path,
    currencies: pd.DataFrame,
    fx: dict[str, pd.DataFrame],
    label: str,
    monthly_fx_fallback: dict[str, pd.DataFrame] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convert a wide monthly file (rows=tickers, cols=YYYY-MM) from local currency
    to NOK using month-end FX.

    Returns:
        (converted_df, audit_df)
        converted_df has the same shape as input, but values are NOK.
        audit_df has a sample of (Ticker, Month, LocalValue, Currency, FXRate, NOKValue)
        for spot-checking.
    """
    df_local = pd.read_csv(local_csv, index_col="Ticker")
    month_cols = [
        c for c in df_local.columns
        if pd.notna(pd.to_datetime(f"{c}-01", errors="coerce")) and str(c) >= OUTPUT_START_MONTH
    ]
    df_local = df_local[month_cols].copy()
    print(f"\nConverting {label}: {df_local.shape[0]} tickers × {df_local.shape[1]} months")

    ccy_map = currencies.set_index("Ticker")["Currency"].to_dict()

    # Pre-compute month-end FX for each currency.
    eom_fx = {}
    for ccy, df_daily in fx.items():
        if ccy == "NOK":
            continue
        if df_daily is None:
            continue
        eom_fx[ccy] = get_month_end_fx(df_daily)
    eom_fx["NOK"] = None  # identity

    converted = df_local.copy().astype(float)
    audit_rows = []
    monthly_fx_fallback = monthly_fx_fallback or {}
    fallback_monthly = {}
    for ccy, df_fallback in monthly_fx_fallback.items():
        fallback_series = df_fallback.set_index("date").sort_index()["rate"]
        fallback_series.index = fallback_series.index.strftime("%Y-%m")
        fallback_monthly[ccy] = fallback_series

    n_no_currency = 0
    n_unsupported_ccy = 0
    n_missing_fx = 0
    n_fallback_fx = 0

    # Determine which (ticker, month) cells need conversion.
    for ticker in df_local.index:
        ccy = ccy_map.get(ticker, None)

        if ccy is None or ccy == "" or pd.isna(ccy):
            converted.loc[ticker] = np.nan
            n_no_currency += 1
            continue

        if ccy == "NOK":
            # No conversion needed; values are already in NOK.
            continue

        if ccy not in eom_fx:
            if ccy not in fallback_monthly:
                converted.loc[ticker] = np.nan
                n_unsupported_ccy += 1
                continue
            rate_series = pd.Series(dtype=float)
        else:
            rate_series = eom_fx[ccy]

        # Multiply each month's value by that month's FX rate.
        for month in df_local.columns:
            local_val = df_local.loc[ticker, month]
            if pd.isna(local_val):
                continue
            if month in rate_series.index:
                rate = rate_series.loc[month]
                if pd.isna(rate):
                    fallback_rate = fallback_monthly.get(ccy, pd.Series(dtype=float)).get(month, np.nan)
                    if pd.notna(fallback_rate):
                        converted.loc[ticker, month] = local_val * fallback_rate
                        n_fallback_fx += 1
                    else:
                        converted.loc[ticker, month] = np.nan
                        n_missing_fx += 1
                else:
                    converted.loc[ticker, month] = local_val * rate
            else:
                fallback_rate = fallback_monthly.get(ccy, pd.Series(dtype=float)).get(month, np.nan)
                if pd.notna(fallback_rate):
                    converted.loc[ticker, month] = local_val * fallback_rate
                    n_fallback_fx += 1
                else:
                    # Month outside FX coverage.
                    converted.loc[ticker, month] = np.nan
                    n_missing_fx += 1

    # Build a small audit sample. For each currency present, take 1-2 sample
    # tickers and 3 sample months (early, mid, late).
    sample_months = []
    if len(df_local.columns) > 0:
        sample_months = [df_local.columns[0]]
        if len(df_local.columns) > 1:
            sample_months.append(df_local.columns[len(df_local.columns) // 2])
        if len(df_local.columns) > 2:
            sample_months.append(df_local.columns[-1])

    seen_ccy = set()
    for ticker in df_local.index:
        ccy = ccy_map.get(ticker, None)
        if ccy is None or pd.isna(ccy):
            continue
        if ccy in seen_ccy and ccy != "NOK":
            # already have a sample for this ccy
            continue
        for month in sample_months:
            local_val = df_local.loc[ticker, month] if month in df_local.columns else np.nan
            if pd.isna(local_val):
                continue
            if ccy == "NOK":
                rate_used = 1.0
                nok_val = local_val
                fx_source = "identity"
            else:
                rate_series = eom_fx.get(ccy)
                rate_used = rate_series.loc[month] if (rate_series is not None and month in rate_series.index) else np.nan
                fx_source = "month_end_norges_bank"
                if pd.isna(rate_used):
                    rate_used = fallback_monthly.get(ccy, pd.Series(dtype=float)).get(month, np.nan)
                    fx_source = "monthly_fx_fallback" if pd.notna(rate_used) else "missing_fx"
                nok_val = local_val * rate_used if pd.notna(rate_used) else np.nan
            audit_rows.append({
                "Source": label,
                "Ticker": ticker,
                "Month": month,
                "Currency": ccy,
                "LocalValue": local_val,
                "FXRate_EoM": rate_used,
                "FXSource_EoM": fx_source,
                "NOKValue": nok_val,
            })
        seen_ccy.add(ccy)
        if len(seen_ccy) >= 8:
            break

    audit_df = pd.DataFrame(audit_rows)

    print(f"  Tickers with no currency:  {n_no_currency}")
    print(f"  Tickers with unsupported currency: {n_unsupported_ccy}")
    print(f"  Cells converted with monthly fallback FX: {n_fallback_fx}")
    print(f"  Cells with missing FX:     {n_missing_fx}")

    return converted, audit_df


def convert_dividends_to_nok(
    long_csv: Path,
    fx: dict[str, pd.DataFrame],
    monthly_fx_fallback: dict[str, pd.DataFrame] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convert long-format dividends to NOK using ex-date FX.

    Required input columns: Ticker, ExDate, DPS_LOCAL, Currency
    Output preserves all input columns and adds DPS_NOK, FXRate, FXDate, FXSource.
    """
    df = pd.read_csv(long_csv)
    print(f"\nConverting dividends: {len(df):,} events")

    required = {"Ticker", "ExDate", "DPS_LOCAL", "Currency"}
    missing_cols = required - set(df.columns)
    if missing_cols:
        raise ValueError(f"{long_csv} is missing required columns: {sorted(missing_cols)}")

    df["ExDate"] = pd.to_datetime(df["ExDate"], errors="coerce")
    df["DPS_LOCAL"] = pd.to_numeric(df["DPS_LOCAL"], errors="coerce")
    df["Currency"] = df["Currency"].astype(str).str.strip().str.upper()
    n_pre_start = (df["ExDate"] < OUTPUT_START_DATE).sum()
    if n_pre_start > 0:
        print(f"  Dropping {n_pre_start:,} dividend rows before {OUTPUT_START_DATE.date()}.")
        df = df.loc[df["ExDate"] >= OUTPUT_START_DATE].copy()
    if "AdjustmentBasis" in df.columns:
        df["AdjustmentBasis"] = df["AdjustmentBasis"].replace({
            "missing_adjustment_factor_default_1": "unverified_missing_adjustment_factor_default_1",
            "factor_field_missing_default_1": "unverified_missing_adjustment_factor_default_1",
        })

    rates = []
    fx_dates = []
    fx_sources = []
    monthly_fx_fallback = monthly_fx_fallback or {}

    for _, row in df.iterrows():
        ccy = row["Currency"]
        ex_date = row["ExDate"]

        if ccy == "NOK":
            rates.append(1.0)
            fx_dates.append(ex_date)
            fx_sources.append("identity")
            continue

        if ccy not in fx or fx[ccy] is None:
            fallback_series = monthly_fx_fallback.get(ccy)
            if fallback_series is not None:
                rate, date_used = get_fallback_rate_for_date(ex_date, fallback_series)
                rates.append(rate)
                fx_dates.append(date_used)
                fx_sources.append("monthly_fx_fallback_prior_month_end" if pd.notna(rate) else "missing_fx_series")
                continue

            rates.append(np.nan)
            fx_dates.append(pd.NaT)
            fx_sources.append("missing_fx_series")
            continue

        rate, date_used = get_rate_for_date(ex_date, fx[ccy])
        if pd.isna(rate):
            fallback_series = monthly_fx_fallback.get(ccy)
            if fallback_series is not None:
                fallback_rate, fallback_date = get_fallback_rate_for_date(ex_date, fallback_series)
                if pd.notna(fallback_rate):
                    rates.append(fallback_rate)
                    fx_dates.append(fallback_date)
                    fx_sources.append("monthly_fx_fallback_prior_month_end")
                    continue

        rates.append(rate)
        fx_dates.append(date_used)
        fx_sources.append("ex_date_or_prior_business_day" if pd.notna(rate) else "no_rate_found")

    df["FXRate"] = rates
    df["FXDate"] = fx_dates
    df["FXSource"] = fx_sources
    df["DPS_NOK"] = df["DPS_LOCAL"] * df["FXRate"]

    n_no_rate = df["FXRate"].isna().sum()
    if n_no_rate > 0:
        print(f"  WARNING: {n_no_rate} dividend rows could not be converted (no FX rate).")

    preferred_cols = [
        "Ticker",
        "ExDate",
        "DPS_UNADJUSTED",
        "DPS_ADJUSTED_GROSS",
        "AdjustmentFactor",
        "DPS_LOCAL",
        "Currency",
        "InstrumentCurrency",
        "AdjustmentBasis",
        "SplitFallbackSource",
        "SplitFallbackEventCount",
        "SplitFallbackEvents",
        "FXRate",
        "FXDate",
        "FXSource",
        "DPS_NOK",
    ]
    ordered_cols = [c for c in preferred_cols if c in df.columns]
    ordered_cols += [c for c in df.columns if c not in ordered_cols]
    out_long = df[ordered_cols].copy()
    return out_long, df


def aggregate_dividends_to_monthly_nok(divs_nok: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate dividend events to monthly NOK DPS, one row per ticker.
    Output shape matches all_stock_prices_nok.csv (rows=tickers, cols=YYYY-MM).
    """
    if divs_nok.empty:
        return pd.DataFrame()

    work = divs_nok.copy()
    work = work.dropna(subset=["DPS_NOK"])
    work["YearMonth"] = pd.to_datetime(work["ExDate"]).dt.to_period("M")

    monthly = (
        work.groupby(["Ticker", "YearMonth"], as_index=False)["DPS_NOK"]
        .sum()
        .pivot(index="Ticker", columns="YearMonth", values="DPS_NOK")
        .fillna(0.0)
    )
    monthly.columns = [str(c) for c in monthly.columns]
    monthly.index.name = "Ticker"
    return monthly


# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 70)
    print("FX CONVERSION: local currency -> NOK")
    print("=" * 70)
    PROCESSED_DATA.mkdir(parents=True, exist_ok=True)
    print(f"Reading local LSEG files from: {RAW_DATA}")
    print(f"Writing NOK outputs to:       {PROCESSED_DATA}")

    # 1. Load currency metadata.
    ccy_path = RAW_DATA / "instrument_currencies.csv"
    if not ccy_path.exists():
        raise FileNotFoundError(f"Run extract_lseg.py first; missing {ccy_path}")

    currencies = pd.read_csv(ccy_path)
    currencies["Ticker"] = currencies["Ticker"].astype(str).str.strip()
    currencies["Currency"] = currencies["Currency"].astype(str).str.strip().str.upper()

    print(f"\nLoaded {len(currencies)} ticker currencies.")
    print(currencies["Currency"].value_counts())

    # 2. Load FX series.
    fx = load_all_fx()
    monthly_fx_fallback = load_monthly_fx_fallback(MONTHLY_FX_FALLBACK)

    needed = set(currencies["Currency"]) - {"NOK", "UNKNOWN", "NAN", ""}
    available_fx = set([c for c in fx if fx[c] is not None]) | set(monthly_fx_fallback)
    missing_fx = needed - available_fx
    if missing_fx:
        print(f"\nWARNING: missing FX series for currencies: {missing_fx}")
        print("Tickers in those currencies will be NaN in NOK output.")

    # 3. Convert prices.
    prices_local_path = RAW_DATA / "all_stock_prices_local.csv"
    if prices_local_path.exists():
        prices_nok, audit_prices = convert_wide_monthly_to_nok(
            prices_local_path, currencies, fx, label="prices", monthly_fx_fallback=monthly_fx_fallback,
        )
        prices_nok.to_csv(PROCESSED_DATA / "all_stock_prices_nok.csv")
        print(f"Saved all_stock_prices_nok.csv: {prices_nok.shape}")
    else:
        print(f"SKIP prices: {prices_local_path} not found")
        audit_prices = pd.DataFrame()

    # 4. Convert market cap.
    mktcap_local_path = RAW_DATA / "historical_market_cap_local.csv"
    if mktcap_local_path.exists():
        mktcap_nok, audit_mktcap = convert_wide_monthly_to_nok(
            mktcap_local_path, currencies, fx, label="market_cap", monthly_fx_fallback=monthly_fx_fallback,
        )
        mktcap_nok.to_csv(PROCESSED_DATA / "historical_market_cap_nok.csv")
        print(f"Saved historical_market_cap_nok.csv: {mktcap_nok.shape}")
    else:
        print(f"SKIP market cap: {mktcap_local_path} not found")
        audit_mktcap = pd.DataFrame()

    # 5. Convert dividends (long format with ex-date FX).
    divs_long_local_path = RAW_DATA / "dividends_raw_long_local.csv"
    if divs_long_local_path.exists():
        divs_long_nok, divs_full = convert_dividends_to_nok(
            divs_long_local_path,
            fx,
            monthly_fx_fallback=monthly_fx_fallback,
        )
        divs_long_nok.to_csv(PROCESSED_DATA / "dividends_raw_long_nok.csv", index=False)
        print(f"Saved dividends_raw_long_nok.csv: {len(divs_long_nok):,} rows")

        # Then aggregate to monthly wide.
        divs_monthly_nok = aggregate_dividends_to_monthly_nok(divs_long_nok)
        divs_monthly_nok.to_csv(PROCESSED_DATA / "dividends_monthly_nok.csv")
        print(f"Saved dividends_monthly_nok.csv: {divs_monthly_nok.shape}")

        # Build a small dividend audit sample.
        audit_divs = divs_long_nok.head(20).copy()
        audit_divs["Source"] = "dividends_sample"
    else:
        print(f"SKIP dividends: {divs_long_local_path} not found")
        audit_divs = pd.DataFrame()

    # 6. CFO forecasts are extracted directly in NOK by extract_lseg.py.
    cfo_raw_path = RAW_DATA / "cfo_forecasts_monthly_raw.csv"
    if cfo_raw_path.exists():
        cfo_out_path = PROCESSED_DATA / "cfo_forecasts_monthly_nok.csv"
        shutil.copyfile(cfo_raw_path, cfo_out_path)
        print(f"Copied CFO forecasts to {cfo_out_path}")
    else:
        print(f"SKIP CFO forecasts: {cfo_raw_path} not found")

    # 7. Combine audit outputs.
    audit_combined = pd.concat(
        [audit_prices, audit_mktcap, audit_divs],
        ignore_index=True,
    )
    audit_path = PROCESSED_DATA / "fx_conversion_audit.csv"
    audit_combined.to_csv(audit_path, index=False)
    print(f"\nSaved audit sample to {audit_path}")
    print("Inspect this file to verify conversions are correct.")
    print("Quick check: AFAGR.HE end-Dec 2024 should show LocalValue ~0.07 EUR")
    print("             FXRate_EoM ~11.79, NOKValue ~0.83.")

    print("\nDone.")


if __name__ == "__main__":
    main()
