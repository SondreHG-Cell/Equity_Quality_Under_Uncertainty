# run_extraction.py

from __future__ import annotations

import argparse
import json
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.simplefilter("ignore", FutureWarning)


# =============================================================================
# Constants
# =============================================================================

EXCHANGES = ["cox", "hex", "isx", "obx", "stx"]
MAX_YEAR_DEFAULT = 2024

PROF_PRIORITY_VARS = {"REVT", "XRD", "BE", "MIB"}
PROF_SUM_VARS = {"COGS", "XSGA_COMPONENTS", "XINT"}
PROF_ALL_VARS = ["REVT", "COGS", "XSGA_COMPONENTS", "XRD", "XINT", "BE", "MIB"]

DA_ALL_VARS = ["COGS_DA", "XSGA_DA"]
DA_SUM_VARS = {"COGS_DA", "XSGA_DA"}

ACC_PRIORITY_VARS = {"ACT", "CHE", "LCT", "TXP", "AT", "OANCF"}
ACC_SUM_VARS = {"STD", "PPEGT"}
ACC_ALL_VARS = ["ACT", "CHE", "LCT", "STD", "TXP", "PPEGT", "AT", "OANCF"]


# =============================================================================
# Path helpers
# =============================================================================

def find_project_root() -> Path:
    """
    Find the real project root by looking for the data subfolders we need.
    """
    here = Path(__file__).resolve().parent if "__file__" in globals() else Path(".").resolve()

    for p in [here] + list(here.parents):
        if (
            (p / "data" / "accounting_sheets").exists()
            and (p / "data" / "mappings").exists()
        ):
            return p

    raise FileNotFoundError(
        "Could not find project root containing data/accounting_sheets and data/mappings."
    )


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def make_extraction_dir(results_root: Path, folder_name: str = "extraction_static") -> Path:
    extraction_dir = results_root / folder_name
    extraction_dir.mkdir(parents=True, exist_ok=True)
    return extraction_dir


def append_log(log_path: Path, message: str) -> None:
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}"
    print(line)
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def write_json(path: Path, obj: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


# =============================================================================
# Generic parsing helpers
# =============================================================================

def parse_year(colname: str) -> Optional[int]:
    if colname is None:
        return None
    s = str(colname).strip()
    m = re.search(r"(19\d{2}|20\d{2})", s)
    if not m:
        return None
    return int(m.group(1))


def to_float(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return np.nan

    s = s.replace("\u00a0", "").replace(" ", "")

    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return np.nan


def clean_label(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = " ".join(s.split())
    return s.lower()


def to_number(x):
    if x is None:
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    if s == "":
        return np.nan

    s = s.replace("\u2212", "-").replace("%", "").replace("\xa0", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")

    try:
        return float(s)
    except Exception:
        return np.nan


# =============================================================================
# Excel sheet helpers
# =============================================================================

def detect_header_row(xlsx_path: Path, sheet_name: str, max_scan_rows: int = 80) -> Optional[int]:
    scan = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        header=None,
        nrows=max_scan_rows,
        usecols="A",
        engine="openpyxl",
    )

    for i in range(len(scan)):
        a0 = scan.iloc[i, 0]
        a0 = "" if pd.isna(a0) else str(a0).strip()
        if a0.lower() == "field name":
            return i

    return None


def read_sheet_all_years(xlsx_path: Path, sheet_name: str) -> Tuple[pd.DataFrame, int]:
    """
    Returns:
        df, fallback_used
    """
    header_row = detect_header_row(xlsx_path, sheet_name)

    if header_row is None:
        df = pd.read_excel(
            xlsx_path,
            sheet_name=sheet_name,
            skiprows=18,
            engine="openpyxl",
        )
        fallback_used = 1
    else:
        df = pd.read_excel(
            xlsx_path,
            sheet_name=sheet_name,
            header=header_row,
            engine="openpyxl",
        )
        fallback_used = 0

    if df.empty or df.shape[1] == 0:
        return df, fallback_used

    label_col = df.columns[0]
    df[label_col] = df[label_col].astype(str).str.strip()
    df = df[df[label_col] != ""].copy()
    return df, fallback_used


def keep_year_cols_upto(df: pd.DataFrame, max_year: int) -> Tuple[pd.DataFrame, List[int]]:
    if df.empty:
        return df, []

    label_col = df.columns[0]
    year_cols = []
    for c in df.columns[1:]:
        y = parse_year(c)
        if y is not None and y <= max_year:
            year_cols.append(c)

    keep_cols = [label_col] + year_cols
    df2 = df[keep_cols].copy()

    new_cols = [label_col] + [parse_year(c) for c in year_cols]
    df2.columns = new_cols

    # Collapse duplicate year columns by first non-null across duplicate columns
    if len(set(new_cols[1:])) != len(new_cols[1:]):
        out = df2[[label_col]].copy()
        for y in sorted(set(new_cols[1:])):
            cols_y = [c for c in df2.columns[1:] if c == y]
            block = df2[cols_y].apply(lambda col: col.map(to_float))
            out[y] = block.bfill(axis=1).iloc[:, 0]
        df2 = out

    years_sorted = sorted([c for c in df2.columns[1:] if isinstance(c, int)])
    return df2, years_sorted


def extract_label_series_with_duplicates(df_years: pd.DataFrame, row_label: str) -> pd.Series:
    if df_years.empty:
        return pd.Series(dtype=float)

    label_col = df_years.columns[0]
    hits = df_years[df_years[label_col] == row_label]
    if hits.empty:
        return pd.Series(dtype=float)

    block = hits.drop(columns=[label_col]).apply(lambda col: col.map(to_float))
    picked = block.bfill(axis=0).iloc[0]
    picked.name = row_label
    return picked


def priority_across_labels(series_list: List[pd.Series]) -> pd.Series:
    if not series_list:
        return pd.Series(dtype=float)
    df = pd.concat(series_list, axis=1)
    return df.bfill(axis=1).iloc[:, 0]


def sum_across_labels(series_list: List[pd.Series]) -> pd.Series:
    if not series_list:
        return pd.Series(dtype=float)
    df = pd.concat(series_list, axis=1)
    return df.sum(axis=1, min_count=1)


def load_mapping(mapping_path: Path) -> Dict:
    return json.loads(mapping_path.read_text(encoding="utf-8"))


# =============================================================================
# Metadata helpers
# =============================================================================

def get_company_metadata(xlsx_path: Path) -> Tuple[str, str]:
    meta = pd.read_excel(
        xlsx_path,
        sheet_name=0,
        header=None,
        usecols="A:B",
        nrows=6,
        engine="openpyxl",
    )

    raw_name = "" if pd.isna(meta.iat[1, 1]) else str(meta.iat[1, 1]).strip()
    industry = "" if pd.isna(meta.iat[4, 1]) else str(meta.iat[4, 1]).strip()

    company_name = re.sub(r"\s*\([^)]*\)\s*$", "", raw_name).strip()
    return company_name, industry


def trbc_industry_to_sector(industry: str) -> str:
    mapping = {
        "Oil & Gas": "Energy",
        "Oil & Gas Related Equipment and Services": "Energy",
        "Renewable Energy": "Energy",
        "Freight & Logistics Services": "Industrials",
        "Machinery, Tools, Heavy Vehicles, Trains & Ships": "Industrials",
        "Construction & Engineering": "Industrials",
        "Professional & Commercial Services": "Industrials",
        "Aerospace & Defense": "Industrials",
        "Passenger Transportation Services": "Industrials",
        "Diversified Industrial Goods Wholesale": "Industrials",
        "Construction Materials": "Industrials",
        "Transport Infrastructure": "Industrials",
        "Food & Tobacco": "Consumer Staples",
        "Beverages": "Consumer Staples",
        "Food & Drug Retailing": "Consumer Staples",
        "Household Goods": "Consumer Staples",
        "Personal & Household Products & Services": "Consumer Staples",
        "Specialty Retailers": "Consumer Discretionary",
        "Diversified Retail": "Consumer Discretionary",
        "Automobiles & Auto Parts": "Consumer Discretionary",
        "Hotels & Entertainment Services": "Consumer Discretionary",
        "Homebuilding & Construction Supplies": "Consumer Discretionary",
        "Textiles & Apparel": "Consumer Discretionary",
        "Leisure Products": "Consumer Discretionary",
        "Consumer Goods Conglomerates": "Consumer Discretionary",
        "Biotechnology & Medical Research": "Health Care",
        "Pharmaceuticals": "Health Care",
        "Healthcare Equipment & Supplies": "Health Care",
        "Healthcare Providers & Services": "Health Care",
        "Banking Services": "Financials",
        "Investment Banking & Investment Services": "Financials",
        "Investment Holding Companies": "Financials",
        "Insurance": "Financials",
        "Metals & Mining": "Materials",
        "Chemicals": "Materials",
        "Paper & Forest Products": "Materials",
        "Containers & Packaging": "Materials",
        "Software & IT Services": "Information Technology",
        "Semiconductors & Semiconductor Equipment": "Information Technology",
        "Electronic Equipment & Parts": "Information Technology",
        "Computers, Phones & Household Electronics": "Information Technology",
        "Integrated Hardware & Software": "Information Technology",
        "Office Equipment": "Information Technology",
        "Communications & Networking": "Communication Services",
        "Media & Publishing": "Communication Services",
        "Telecommunications Services": "Communication Services",
        "Electric Utilities & IPPs": "Utilities",
        "Real Estate Operations": "Real Estate",
        "Professional & Business Education": "Industrials",
        "Miscellaneous Educational Service Providers": "Industrials",
    }

    if industry is None:
        return "Other"
    return mapping.get(str(industry).strip(), "Other")


# =============================================================================
# Generic mapped extraction
# =============================================================================

def extract_mapped_variables(
    xlsx_path: Path,
    mapping_path: Path,
    all_vars: List[str],
    priority_vars: set[str],
    sum_vars: set[str],
    max_year: int,
) -> Tuple[pd.DataFrame, int]:
    mapping = load_mapping(mapping_path)
    variables = mapping.get("variables", [])

    sheet_cache: Dict[str, pd.DataFrame] = {}
    years_union = set()
    extracted: Dict[str, pd.Series] = {}
    fallback_count = 0

    for var_item in variables:
        var = var_item.get("variable", "")
        if var not in all_vars:
            continue

        choices = var_item.get("final_choice", [])
        if not choices:
            extracted[var] = pd.Series(dtype=float)
            continue

        label_series_list: List[pd.Series] = []

        for ch in choices:
            sheet = ch["sheet_name"]
            label = ch["row_label"]

            if sheet not in sheet_cache:
                df_raw, n_fb = read_sheet_all_years(xlsx_path, sheet)
                fallback_count += n_fb
                df_years, _ = keep_year_cols_upto(df_raw, max_year)
                sheet_cache[sheet] = df_years

            s = extract_label_series_with_duplicates(sheet_cache[sheet], label)
            if not s.empty:
                label_series_list.append(s)

        if var in sum_vars:
            s_var = sum_across_labels(label_series_list)
        else:
            s_var = priority_across_labels(label_series_list)

        extracted[var] = s_var
        years_union.update(list(s_var.index))

    years_sorted = sorted([y for y in years_union if isinstance(y, int)])
    df_out = pd.DataFrame(index=years_sorted)
    df_out.index.name = "Year"

    for var in all_vars:
        s = extracted.get(var, pd.Series(dtype=float))
        df_out[var] = s.reindex(years_sorted)

    df_out = df_out.fillna(0.0)
    return df_out.reset_index(), fallback_count


# =============================================================================
# Market cap extraction
# =============================================================================

def extract_market_cap_from_valuation(xlsx_path: Path) -> pd.DataFrame:
    """
    Return DataFrame with columns: Year, MarketCap
    from the 'Valuation' sheet.

    Picks the row named 'Market Capitalization' that actually has values,
    not the header/section row.

    If duplicate year columns exist in the sheet, collapse them to one row
    per Year by keeping the first non-null value from left to right.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    if "Valuation" not in wb.sheetnames:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    ws = wb["Valuation"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    # Find year header row
    year_row_idx = None
    year_cols = []

    for i, row in enumerate(rows):
        candidate_cols = []
        for j, val in enumerate(row):
            if val is None:
                continue
            sval = str(val).strip()
            if sval.isdigit() and len(sval) == 4:
                yr = int(sval)
                if 1900 <= yr <= 2100:
                    candidate_cols.append(j)

        if len(candidate_cols) >= 2:
            year_row_idx = i
            year_cols = candidate_cols
            break

    if year_row_idx is None or not year_cols:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    years = {}
    for j in year_cols:
        val = rows[year_row_idx][j]
        try:
            years[j] = int(str(val).strip())
        except Exception:
            pass

    if not years:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    # Find correct Market Capitalization row
    chosen_row = None
    for row in rows:
        label = clean_label(row[0] if len(row) > 0 else "")
        if label != "market capitalization":
            continue

        numeric_count = 0
        for j in years.keys():
            val = row[j] if j < len(row) else None
            num = to_number(val)
            if pd.notna(num):
                numeric_count += 1

        if numeric_count >= 1:
            chosen_row = row
            break

    if chosen_row is None:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    out = []
    for j, yr in years.items():
        val = chosen_row[j] if j < len(chosen_row) else None
        num = to_number(val)
        if pd.notna(num):
            out.append(
                {
                    "Year": int(yr),
                    "MarketCap": float(num),
                    "_col_order": int(j),   # preserve left-to-right order
                }
            )

    if not out:
        return pd.DataFrame(columns=["Year", "MarketCap"])

    mc_df = pd.DataFrame(out).sort_values(["Year", "_col_order"]).reset_index(drop=True)

    # Collapse duplicate years by taking the first non-null value from left to right
    mc_df = (
        mc_df.groupby("Year", as_index=False)
        .agg(MarketCap=("MarketCap", "first"))
        .sort_values("Year")
        .reset_index(drop=True)
    )

    return mc_df


# =============================================================================
# Firm-level extraction
# =============================================================================

def build_prof_firm_output(
    xlsx_path: Path,
    prof_mapping_path: Path,
    da_mapping_path: Optional[Path],
    max_year: int,
) -> Tuple[pd.DataFrame, dict]:
    prof_df, prof_fallbacks = extract_mapped_variables(
        xlsx_path=xlsx_path,
        mapping_path=prof_mapping_path,
        all_vars=PROF_ALL_VARS,
        priority_vars=PROF_PRIORITY_VARS,
        sum_vars=PROF_SUM_VARS,
        max_year=max_year,
    )

    da_fallbacks = 0
    if da_mapping_path is not None and da_mapping_path.exists():
        da_df, da_fallbacks = extract_mapped_variables(
            xlsx_path=xlsx_path,
            mapping_path=da_mapping_path,
            all_vars=DA_ALL_VARS,
            priority_vars=set(),
            sum_vars=DA_SUM_VARS,
            max_year=max_year,
        )
        prof_df = prof_df.merge(da_df, on="Year", how="left")
    else:
        prof_df["COGS_DA"] = 0.0
        prof_df["XSGA_DA"] = 0.0

    for c in PROF_ALL_VARS + DA_ALL_VARS:
        if c in prof_df.columns:
            prof_df[c] = pd.to_numeric(prof_df[c], errors="coerce").fillna(0.0)

    prof_df["COGS"] = prof_df["COGS"] + prof_df["COGS_DA"]
    prof_df["XSGA_COMPONENTS"] = prof_df["XSGA_COMPONENTS"] + prof_df["XSGA_DA"]

    # Expense-like items positive
    for c in ["REVT", "COGS", "XSGA_COMPONENTS", "XRD", "XINT"]:
        if c in prof_df.columns:
            prof_df[c] = prof_df[c].abs()

    numer = prof_df["REVT"] - prof_df["COGS"] - (prof_df["XSGA_COMPONENTS"] - prof_df["XRD"]) - prof_df["XINT"]
    denom = (prof_df["BE"] + prof_df["MIB"]).replace({0.0: np.nan})
    prof_df["PROF"] = numer / denom

    mc_df = extract_market_cap_from_valuation(xlsx_path)

    if not mc_df.empty:
        mc_df = (
            mc_df.sort_values("Year")
            .drop_duplicates(subset=["Year"], keep="first")
            .reset_index(drop=True)
        )
        prof_df = prof_df.merge(mc_df, on="Year", how="left", validate="1:1")
    else:
        prof_df["MarketCap"] = np.nan

    company_name, industry = get_company_metadata(xlsx_path)
    sector = trbc_industry_to_sector(industry)
    ticker = xlsx_path.stem

    prof_df.insert(1, "Ticker", ticker)
    prof_df.insert(2, "CompanyName", company_name)
    prof_df.insert(3, "Industry", industry)
    prof_df.insert(4, "Sector", sector)

    prof_df = prof_df[
        [
            "Year",
            "Ticker",
            "CompanyName",
            "Industry",
            "Sector",
            "REVT",
            "COGS",
            "XSGA_COMPONENTS",
            "XRD",
            "XINT",
            "BE",
            "MIB",
            "MarketCap",
            "PROF",
        ]
    ].copy()

    stats = {
        "prof_fallbacks": prof_fallbacks,
        "da_fallbacks": da_fallbacks,
        "marketcap_nonnull": int(prof_df["MarketCap"].notna().sum()),
        "n_rows": int(len(prof_df)),
    }

    return prof_df, stats


def build_acc_firm_output(
    xlsx_path: Path,
    acc_mapping_path: Path,
    max_year: int,
) -> Tuple[pd.DataFrame, dict]:
    acc_df, acc_fallbacks = extract_mapped_variables(
        xlsx_path=xlsx_path,
        mapping_path=acc_mapping_path,
        all_vars=ACC_ALL_VARS,
        priority_vars=ACC_PRIORITY_VARS,
        sum_vars=ACC_SUM_VARS,
        max_year=max_year,
    )

    company_name, industry = get_company_metadata(xlsx_path)
    sector = trbc_industry_to_sector(industry)
    ticker = xlsx_path.stem

    acc_df.insert(1, "Ticker", ticker)
    acc_df.insert(2, "CompanyName", company_name)
    acc_df.insert(3, "Industry", industry)
    acc_df.insert(4, "Sector", sector)

    acc_df = acc_df[
        [
            "Year",
            "Ticker",
            "CompanyName",
            "Industry",
            "Sector",
            "ACT",
            "CHE",
            "LCT",
            "STD",
            "TXP",
            "PPEGT",
            "AT",
            "OANCF",
        ]
    ].copy()

    stats = {
        "acc_fallbacks": acc_fallbacks,
        "n_rows": int(len(acc_df)),
    }

    return acc_df, stats


# =============================================================================
# Exchange-level extraction
# =============================================================================

def process_prof_exchange(
    exchange: str,
    xlsx_dir: Path,
    prof_mappings_dir: Path,
    da_mappings_dir: Path,
    prof_out_dir: Path,
    max_year: int,
    log_path: Path,
) -> dict:
    mapping_files = sorted(prof_mappings_dir.glob("*.json"))
    append_log(log_path, f"[{exchange.upper()}] PROF mappings found: {len(mapping_files)}")

    stats = {
        "exchange": exchange,
        "n_mappings": len(mapping_files),
        "n_extracted": 0,
        "n_missing_xlsx": 0,
        "n_errors": 0,
        "prof_fallbacks": 0,
        "da_fallbacks": 0,
        "marketcap_rows_nonnull": 0,
    }

    for mp in mapping_files:
        ticker = mp.stem
        xlsx_path = xlsx_dir / f"{ticker}.xlsx"
        da_mapping_path = da_mappings_dir / f"{ticker}.json"

        if not xlsx_path.exists():
            stats["n_missing_xlsx"] += 1
            continue

        try:
            firm_df, firm_stats = build_prof_firm_output(
                xlsx_path=xlsx_path,
                prof_mapping_path=mp,
                da_mapping_path=da_mapping_path,
                max_year=max_year,
            )

            firm_df.to_csv(prof_out_dir / f"{ticker}.csv", index=False)

            stats["n_extracted"] += 1
            stats["prof_fallbacks"] += firm_stats["prof_fallbacks"]
            stats["da_fallbacks"] += firm_stats["da_fallbacks"]
            stats["marketcap_rows_nonnull"] += firm_stats["marketcap_nonnull"]

        except Exception as e:
            append_log(log_path, f"[{exchange.upper()}][PROF][ERROR] {ticker}: {type(e).__name__}: {e}")
            stats["n_errors"] += 1

    append_log(
        log_path,
        f"[{exchange.upper()}][PROF] extracted={stats['n_extracted']}, "
        f"missing_xlsx={stats['n_missing_xlsx']}, errors={stats['n_errors']}, "
        f"prof_fallbacks={stats['prof_fallbacks']}, da_fallbacks={stats['da_fallbacks']}",
    )
    return stats


def process_acc_exchange(
    exchange: str,
    xlsx_dir: Path,
    acc_mappings_dir: Path,
    acc_out_dir: Path,
    max_year: int,
    log_path: Path,
) -> dict:
    mapping_files = sorted(acc_mappings_dir.glob("*.json"))
    append_log(log_path, f"[{exchange.upper()}] ACC mappings found: {len(mapping_files)}")

    stats = {
        "exchange": exchange,
        "n_mappings": len(mapping_files),
        "n_extracted": 0,
        "n_missing_xlsx": 0,
        "n_errors": 0,
        "acc_fallbacks": 0,
    }

    for mp in mapping_files:
        ticker = mp.stem
        xlsx_path = xlsx_dir / f"{ticker}.xlsx"

        if not xlsx_path.exists():
            stats["n_missing_xlsx"] += 1
            continue

        try:
            firm_df, firm_stats = build_acc_firm_output(
                xlsx_path=xlsx_path,
                acc_mapping_path=mp,
                max_year=max_year,
            )

            firm_df.to_csv(acc_out_dir / f"{ticker}.csv", index=False)

            stats["n_extracted"] += 1
            stats["acc_fallbacks"] += firm_stats["acc_fallbacks"]

        except Exception as e:
            append_log(log_path, f"[{exchange.upper()}][ACC][ERROR] {ticker}: {type(e).__name__}: {e}")
            stats["n_errors"] += 1

    append_log(
        log_path,
        f"[{exchange.upper()}][ACC] extracted={stats['n_extracted']}, "
        f"missing_xlsx={stats['n_missing_xlsx']}, errors={stats['n_errors']}, "
        f"acc_fallbacks={stats['acc_fallbacks']}",
    )
    return stats


# =============================================================================
# Build prepared Step 2 input
# =============================================================================

def load_all_csvs(folder: Path) -> pd.DataFrame:
    files = sorted(folder.glob("*.csv"))
    if not files:
        raise ValueError(f"No CSV files found in {folder}")

    frames = []
    for fp in files:
        df = pd.read_csv(fp)
        frames.append(df)

    out = pd.concat(frames, ignore_index=True)
    out["Year"] = pd.to_numeric(out["Year"], errors="coerce")
    out = out.dropna(subset=["Ticker", "Year"]).copy()
    out["Year"] = out["Year"].astype(int)
    out = out.sort_values(["Ticker", "Year"]).reset_index(drop=True)
    return out


def build_prepared_step2_input(
    acc_out_dir: Path,
    prof_out_dir: Path,
) -> pd.DataFrame:
    acc_panel = load_all_csvs(acc_out_dir)
    prof_panel = load_all_csvs(prof_out_dir)

    prof_keep = [
        "Ticker",
        "Year",
        "REVT",
        "COGS",
        "XSGA_COMPONENTS",
        "XRD",
        "XINT",
        "BE",
        "MIB",
        "MarketCap",
        "PROF",
    ]

    # Keep only needed columns on PROF side
    prof_panel = prof_panel[prof_keep].copy()

    # --------------------------------------------------
    # Drop exact duplicate rows first
    # --------------------------------------------------
    acc_before = len(acc_panel)
    prof_before = len(prof_panel)

    acc_panel = acc_panel.drop_duplicates().copy()
    prof_panel = prof_panel.drop_duplicates().copy()

    acc_dropped = acc_before - len(acc_panel)
    prof_dropped = prof_before - len(prof_panel)

    if acc_dropped > 0:
        print(f"[build_prepared_step2_input] Dropped {acc_dropped} exact duplicate rows from acc_panel")
    if prof_dropped > 0:
        print(f"[build_prepared_step2_input] Dropped {prof_dropped} exact duplicate rows from prof_panel")

    # --------------------------------------------------
    # Check key uniqueness
    # --------------------------------------------------
    acc_key_dupes = acc_panel.duplicated(subset=["Ticker", "Year"], keep=False)
    prof_key_dupes = prof_panel.duplicated(subset=["Ticker", "Year"], keep=False)

    if acc_key_dupes.any():
        dupes = (
            acc_panel.loc[acc_key_dupes, ["Ticker", "Year"]]
            .value_counts()
            .reset_index(name="n_rows")
            .sort_values(["n_rows", "Ticker", "Year"], ascending=[False, True, True])
        )
        print("\n[build_prepared_step2_input] Duplicate keys found in acc_panel:")
        print(dupes.head(20).to_string(index=False))

        raise ValueError(
            "acc_panel still has duplicate (Ticker, Year) keys after dropping exact duplicates. "
            "Inspect the printed sample above."
        )

    if prof_key_dupes.any():
        dupes = (
            prof_panel.loc[prof_key_dupes, ["Ticker", "Year"]]
            .value_counts()
            .reset_index(name="n_rows")
            .sort_values(["n_rows", "Ticker", "Year"], ascending=[False, True, True])
        )
        print("\n[build_prepared_step2_input] Duplicate keys found in prof_panel:")
        print(dupes.head(20).to_string(index=False))

        # Also show the full duplicated rows for first few duplicate keys
        first_keys = dupes[["Ticker", "Year"]].head(5)
        sample = prof_panel.merge(first_keys, on=["Ticker", "Year"], how="inner")
        print("\n[build_prepared_step2_input] Sample duplicated PROF rows:")
        print(sample.sort_values(["Ticker", "Year"]).to_string(index=False))

        raise ValueError(
            "prof_panel still has duplicate (Ticker, Year) keys after dropping exact duplicates. "
            "Inspect the printed sample above."
        )

    # --------------------------------------------------
    # Merge
    # --------------------------------------------------
    prepared = acc_panel.merge(
        prof_panel,
        on=["Ticker", "Year"],
        how="inner",
        validate="1:1",
    ).copy()

    first_cols = [
        "Year",
        "Ticker",
        "CompanyName",
        "Industry",
        "Sector",
        "ACT",
        "CHE",
        "LCT",
        "STD",
        "TXP",
        "PPEGT",
        "AT",
        "OANCF",
        "REVT",
        "COGS",
        "XSGA_COMPONENTS",
        "XRD",
        "XINT",
        "BE",
        "MIB",
        "MarketCap",
        "PROF",
    ]
    existing_first = [c for c in first_cols if c in prepared.columns]
    remaining = [c for c in prepared.columns if c not in existing_first]
    prepared = prepared[existing_first + remaining]

    return prepared.sort_values(["Ticker", "Year"]).reset_index(drop=True)


# =============================================================================
# Main run function
# =============================================================================

def run_extraction(
    results_root: str | Path = "results",
    run_name: Optional[str] = None,
    exchanges: Optional[List[str]] = None,
    max_year: int = MAX_YEAR_DEFAULT,
) -> dict:
    project_root = find_project_root()
    data_dir = project_root / "data"

    accounting_sheets_dir = data_dir / "accounting_sheets"
    prof_mappings_root = data_dir / "mappings" / "prof_mappings"
    acc_mappings_root = data_dir / "mappings" / "acc_mappings"
    da_mappings_root = data_dir / "mappings" / "da_mappings"
    stock_prices_csv = data_dir / "all_stock_prices.csv"

    results_root = Path(results_root)
    if not results_root.is_absolute():
        results_root = project_root / results_root
    results_root.mkdir(parents=True, exist_ok=True)

    extraction_dir = make_extraction_dir(results_root, folder_name="extraction_static")

    prof_out_dir = extraction_dir / "prof_components_extracted"
    acc_out_dir = extraction_dir / "acc_components_extracted"
    prof_out_dir.mkdir(parents=True, exist_ok=True)
    acc_out_dir.mkdir(parents=True, exist_ok=True)

    for fp in prof_out_dir.glob("*.csv"):
        fp.unlink()

    for fp in acc_out_dir.glob("*.csv"):
        fp.unlink()
    
    log_path = extraction_dir / "extraction_log.txt"
    config_path = extraction_dir / "extraction_config.json"
    summary_path = extraction_dir / "extraction_summary.json"

    exchanges = exchanges or EXCHANGES

    config = {
        "project_root": str(project_root),
        "results_root": str(results_root),
        "extraction_dir": str(extraction_dir),
        "exchanges": exchanges,
        "max_year": max_year,
        "stock_prices_csv": str(stock_prices_csv),
    }
    write_json(config_path, config)

    append_log(log_path, f"Created extraction run directory: {extraction_dir}")

    prof_exchange_stats = []
    acc_exchange_stats = []

    for ex in exchanges:
        xlsx_dir = accounting_sheets_dir / f"{ex}_xlsx"
        prof_mappings_dir = prof_mappings_root / f"mappings_{ex}"
        acc_mappings_dir = acc_mappings_root / f"acc_mappings_{ex}"
        da_mappings_dir = da_mappings_root / f"da_mappings_{ex}"

        if not xlsx_dir.exists():
            append_log(log_path, f"[{ex.upper()}] missing xlsx dir: {xlsx_dir}")
            continue
        if not prof_mappings_dir.exists():
            append_log(log_path, f"[{ex.upper()}] missing prof mappings dir: {prof_mappings_dir}")
            continue
        if not acc_mappings_dir.exists():
            append_log(log_path, f"[{ex.upper()}] missing acc mappings dir: {acc_mappings_dir}")
            continue
        if not da_mappings_dir.exists():
            append_log(log_path, f"[{ex.upper()}] missing da mappings dir: {da_mappings_dir}")

        prof_stats = process_prof_exchange(
            exchange=ex,
            xlsx_dir=xlsx_dir,
            prof_mappings_dir=prof_mappings_dir,
            da_mappings_dir=da_mappings_dir,
            prof_out_dir=prof_out_dir,
            max_year=max_year,
            log_path=log_path,
        )
        prof_exchange_stats.append(prof_stats)

        acc_stats = process_acc_exchange(
            exchange=ex,
            xlsx_dir=xlsx_dir,
            acc_mappings_dir=acc_mappings_dir,
            acc_out_dir=acc_out_dir,
            max_year=max_year,
            log_path=log_path,
        )
        acc_exchange_stats.append(acc_stats)

    append_log(log_path, "Building prepared Step 2 input panel")
    prepared_df = build_prepared_step2_input(acc_out_dir=acc_out_dir, prof_out_dir=prof_out_dir)

    prepared_input_csv = extraction_dir / "prepared_step2_input.csv"
    prepared_df.to_csv(prepared_input_csv, index=False)

    append_log(log_path, f"Saved prepared Step 2 input: {prepared_input_csv}")
    append_log(log_path, f"Prepared input shape: {prepared_df.shape}")

    summary = {
        "extraction_dir": str(extraction_dir),
        "prof_components_dir": str(prof_out_dir),
        "acc_components_dir": str(acc_out_dir),
        "prepared_step2_input_csv": str(prepared_input_csv),
        "stock_prices_csv": str(stock_prices_csv),
        "prepared_rows": int(len(prepared_df)),
        "prepared_unique_tickers": int(prepared_df["Ticker"].nunique()),
        "prepared_year_min": int(prepared_df["Year"].min()) if not prepared_df.empty else None,
        "prepared_year_max": int(prepared_df["Year"].max()) if not prepared_df.empty else None,
        "marketcap_nonnull_rows": int(prepared_df["MarketCap"].notna().sum()) if "MarketCap" in prepared_df.columns else 0,
        "prof_exchange_stats": prof_exchange_stats,
        "acc_exchange_stats": acc_exchange_stats,
    }
    write_json(summary_path, summary)

    append_log(log_path, "Extraction run completed successfully.")

    return {
        "extraction_dir": str(extraction_dir),
        "prof_components_dir": str(prof_out_dir),
        "acc_components_dir": str(acc_out_dir),
        "prepared_step2_input_csv": str(prepared_input_csv),
        "stock_prices_csv": str(stock_prices_csv),
        "extraction_config_json": str(config_path),
        "extraction_summary_json": str(summary_path),
        "extraction_log_txt": str(log_path),
    }


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run full extraction pipeline.")

    parser.add_argument(
        "--results_root",
        type=str,
        default="results",
        help="Root folder for timestamped extraction runs.",
    )
    parser.add_argument(
        "--run_name",
        type=str,
        default=None,
        help="Optional run name suffix.",
    )
    parser.add_argument(
        "--max_year",
        type=int,
        default=MAX_YEAR_DEFAULT,
        help="Maximum year to keep from extracted sheets.",
    )
    parser.add_argument(
        "--exchanges",
        nargs="*",
        default=EXCHANGES,
        help="Exchanges to process. Example: --exchanges obx stx",
    )

    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    result = run_extraction(
        results_root=args.results_root,
        run_name=args.run_name,
        exchanges=args.exchanges,
        max_year=args.max_year,
    )

    print("\nSaved extraction outputs:")
    for key, value in result.items():
        print(f"  {key}: {value}")